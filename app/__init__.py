from flask import Flask, render_template, request, send_file, redirect, url_for, flash, jsonify
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.utils import range_boundaries
import io
import os

def create_app():
    app = Flask(__name__)
    app.secret_key = "change-this-key"

    # ---- Paths (files live in /app alongside this __init__.py) ----
    BASE_DIR = os.path.dirname(__file__)
    TEMPLATE_PATH = os.path.join(BASE_DIR, "AIM_Change_Request_Template.xlsx")
    MASTERLIST_PATH = os.path.join(BASE_DIR, "masterlist.xlsx")

    # ---- Template config ----
    TARGET_SHEET_NAME = os.environ.get("TARGET_SHEET_NAME", "Sheet1")
    START_ROW = int(os.environ.get("START_ROW", "32"))

    COLUMN_MAP = {
        "wrin": "A",
        "description": "B",      # B:C will be merged per data row
        "reduction_type": "D",
        "stock_on_hand": "E",    # allows decimals
        "reason": "F",
        "comment": "G",
    }

    # Header cells (top-left cell of merged areas)
    HEADER_CELLS = {
        "restaurant_name": "C25",
        "restaurant_number": "C26",
        "manager_name_cell": "C27",
        "current_date_cell": "C28",
        "delivery_date_cell": "C29",
    }

    REDUCTION_CHOICES = ["Increase", "Decrease"]
    REASON_CHOICES = [
        "Stock on Hand Variance",
        "Manual Items",
        "Safety Stock",
        "Shelf Life",
        "Usage",
    ]

    RESTAURANTS = [
        {"name": "Tugun",           "number": "1814"},
        {"name": "Nerang",          "number": "253"},
        {"name": "Robina Central",  "number": "786"},
        {"name": "Robina FC II",    "number": "1220"},
        {"name": "Reedy Creek",     "number": "886"},
        {"name": "Burleigh Waters", "number": "395"},
        {"name": "Palm Beach",      "number": "1047"},
        {"name": "Elanora",         "number": "393"},
    ]

    # ---- Masterlist (WRIN <-> Name) ----
    MASTER_ROWS = []      # [{wrin, name}]
    WRIN_TO_NAME = {}     # "12345" -> "Widget"
    NAME_TO_WRIN = {}     # "widget name".lower() -> "12345"

    def load_masterlist():
        if not os.path.exists(MASTERLIST_PATH):
            return
        try:
            wb = load_workbook(MASTERLIST_PATH, data_only=True, read_only=True)
            ws = wb.active  # assumes first sheet: col A=WRIN, col B=Name
            for r in ws.iter_rows(min_row=2, max_col=2, values_only=True):
                w, n = r[0], r[1]
                if w is None or n is None:
                    continue
                wrin = str(w).strip()
                name = str(n).strip()
                if not wrin or not name:
                    continue
                MASTER_ROWS.append({"wrin": wrin, "name": name})
                WRIN_TO_NAME[wrin] = name
                NAME_TO_WRIN[name.lower()] = wrin
        except Exception:
            pass

    load_masterlist()

    # ---- Helpers ----
    def row_merge_exists(ws, rng: str) -> bool:
        return any(str(r) == rng for r in ws.merged_cells.ranges)

    def ensure_merge(ws, start_col_letter: str, end_col_letter: str, row: int):
        """Merge e.g. B{row}:C{row}, skip if template already merged."""
        rng = f"{start_col_letter}{row}:{end_col_letter}{row}"
        if row_merge_exists(ws, rng):
            return
        sc, sr, ec, er = range_boundaries(rng)
        for mr in ws.merged_cells.ranges:
            msc, msr, mec, mer = mr.bounds
            if not (er < msr or sr > mer or ec < msc or sc > mec):
                return
        ws.merge_cells(rng)

    def to_number_or_keep(x: str):
        x = (x or "").strip()
        if x == "":
            return ""
        try:
            return float(x)  # allow decimals
        except ValueError:
            return x

    def ymd_to_dmy(ymd: str) -> str:
        """
        Convert 'YYYY-MM-DD' to 'DD/MM/YYYY'.
        If blank or invalid, return as-is.
        """
        ymd = (ymd or "").strip()
        try:
            dt = datetime.strptime(ymd, "%Y-%m-%d")
            return dt.strftime("%d/%m/%Y")
        except ValueError:
            return ymd

    # ---- Routes ----
    @app.route("/", methods=["GET"])
    def index():
        restaurant_options = [f"{r['name']} — {r['number']}" for r in RESTAURANTS]
        return render_template(
            "index.html",
            reduction_choices=REDUCTION_CHOICES,
            reason_choices=REASON_CHOICES,
            restaurant_options=restaurant_options,
        )

    @app.route("/lookup", methods=["GET"])
    def lookup():
        q = (request.args.get("q") or "").strip()
        mode = (request.args.get("mode") or "wrin").strip().lower()
        if not q or not MASTER_ROWS:
            return jsonify([])

        qlow = q.lower()
        if mode == "name":
            prefix = [r for r in MASTER_ROWS if r["name"].lower().startswith(qlow)]
            substr = [r for r in MASTER_ROWS if qlow in r["name"].lower() and r not in prefix]
        else:  # wrin
            prefix = [r for r in MASTER_ROWS if r["wrin"].lower().startswith(qlow)]
            substr = [r for r in MASTER_ROWS if qlow in r["wrin"].lower() and r not in prefix]

        return jsonify((prefix + substr)[:15])

    @app.route("/generate", methods=["POST"])
    def generate():
        # Header fields
        restaurant_choice = (request.form.get("restaurant") or "").strip()
        delivery_date_raw = (request.form.get("delivery_date") or "").strip()  # 'YYYY-MM-DD' from <input type="date">
        manager_name = (request.form.get("manager_name") or "").strip()

        # Current date in dd/mm/yyyy for Excel
        current_date_dmy = datetime.now().strftime("%d/%m/%Y")
        # Delivery date converted to dd/mm/yyyy for Excel
        delivery_date_dmy = ymd_to_dmy(delivery_date_raw)

        # Rows
        wrins = request.form.getlist("wrin[]")
        descriptions = request.form.getlist("description[]")
        reductions = request.form.getlist("reduction_type[]")
        onhands = request.form.getlist("stock_on_hand[]")
        reasons = request.form.getlist("reason[]")
        comments = request.form.getlist("comment[]")

        if not wrins:
            flash("Please add at least one item.", "error")
            return redirect(url_for("index"))

        if not os.path.exists(TEMPLATE_PATH):
            flash(f"Template not found at: {TEMPLATE_PATH}", "error")
            return redirect(url_for("index"))

        try:
            wb = load_workbook(TEMPLATE_PATH)
        except Exception as e:
            flash(f"Could not open template: {e}", "error")
            return redirect(url_for("index"))

        if TARGET_SHEET_NAME not in wb.sheetnames:
            flash(f"Sheet '{TARGET_SHEET_NAME}' not found in template.", "error")
            return redirect(url_for("index"))

        ws = wb[TARGET_SHEET_NAME]

        # Restaurant selection
        selected = None
        if restaurant_choice:
            for r in RESTAURANTS:
                if f"{r['name']} — {r['number']}" == restaurant_choice:
                    selected = r
                    break

        # Write header cells (dd/mm/yyyy for dates)
        if selected:
            ws[HEADER_CELLS["restaurant_name"]] = selected["name"]
            ws[HEADER_CELLS["restaurant_number"]] = selected["number"]
        if HEADER_CELLS.get("manager_name_cell"):
            ws[HEADER_CELLS["manager_name_cell"]] = manager_name
        if HEADER_CELLS.get("current_date_cell"):
            ws[HEADER_CELLS["current_date_cell"]] = current_date_dmy
        if HEADER_CELLS.get("delivery_date_cell"):
            ws[HEADER_CELLS["delivery_date_cell"]] = delivery_date_dmy

        # Write rows
        write_row = START_ROW
        for i in range(len(wrins)):
            wrin = (wrins[i] or "").strip()
            desc = (descriptions[i] or "").strip()
            red = (reductions[i] or "").strip()
            soh = to_number_or_keep(onhands[i] if i < len(onhands) else "")
            reason = (reasons[i] or "").strip()
            comment = (comments[i] or "").strip()

            # Fill missing pair from master on exact match
            if wrin and not desc and wrin in WRIN_TO_NAME:
                desc = WRIN_TO_NAME[wrin]
            if desc and not wrin:
                w_guess = NAME_TO_WRIN.get(desc.lower())
                if w_guess:
                    wrin = w_guess

            ws[f"{COLUMN_MAP['wrin']}{write_row}"] = wrin
            ws[f"{COLUMN_MAP['description']}{write_row}"] = desc
            ws[f"{COLUMN_MAP['reduction_type']}{write_row}"] = red
            ws[f"{COLUMN_MAP['stock_on_hand']}{write_row}"] = soh
            ws[f"{COLUMN_MAP['reason']}{write_row}"] = reason
            ws[f"{COLUMN_MAP['comment']}{write_row}"] = comment

            ensure_merge(ws, "B", "C", write_row)
            write_row += 1

        # Return the filled workbook
        out = io.BytesIO()
        wb.save(out)
        out.seek(0)
        # Use ddmmyyyy in filename too (nice touch)
        fname = f"AimOrderAmendment_{datetime.now().strftime('%d%m%Y')}.xlsx"
        return send_file(
            out,
            as_attachment=True,
            download_name=fname,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    return app

# For PyCharm run or `flask run`
app = create_app()
